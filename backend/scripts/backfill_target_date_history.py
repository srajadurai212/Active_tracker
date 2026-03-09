"""
Backfill target_date_history from Excel file.

For each activity row in the Excel that has a multiline target closure date
(meaning the date was revised one or more times), this script:
  1. Parses the date sequence (newest first in Excel → reversed to oldest first)
  2. Updates initial_target_closure_date to the oldest date
  3. Inserts TargetDateHistory rows for each revision step
  4. Also inserts an initial entry for activities with a single target date
     (where initial_target_closure_date was already set but no history row exists)

Run from backend/ directory:
    python3 scripts/backfill_target_date_history.py
"""
from __future__ import annotations
import asyncio
import sys
import re
import uuid
from datetime import datetime, date, timedelta, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker
from sqlalchemy import select, func

from app.db.models import Activity, TargetDateHistory, User

DATABASE_URL = "postgresql+asyncpg://activitytracker_dev_user:activity4dev18762@157.173.220.118:10000/activitytracker"
EXCEL_PATH = Path(__file__).parent.parent.parent.parent / "IZ - Leads Activity Tracker.xlsx"


def parse_date(s: str) -> date | None:
    s = s.strip()
    if not s:
        return None
    # Fix common typos
    s = s.replace("_", "-")
    # Replace full month names with abbreviations
    month_map = {
        "January": "Jan", "February": "Feb", "March": "Mar", "April": "Apr",
        "May": "May", "June": "Jun", "July": "Jul", "August": "Aug",
        "September": "Sep", "October": "Oct", "November": "Nov", "December": "Dec",
    }
    for full, abbr in month_map.items():
        s = s.replace(full, abbr)

    formats = [
        "%d-%b-%y",    # 16-Apr-25
        "%d-%b-%Y",    # 16-Apr-2025
        "%d-%m-%Y",    # 18-06-2025
        "%m/%d/%Y",    # 7/25/2025
        "%d/%m/%Y",    # 25/06/2025
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    print(f"  WARNING: Could not parse date: {repr(s)}")
    return None


def normalize_str(s: str) -> str:
    """Lowercase + collapse whitespace for fuzzy matching."""
    return re.sub(r"\s+", " ", (s or "").lower().strip())


async def main() -> None:
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Consolidated Activity Tracker"]

    # ── Build Excel lookup: (client_norm, entry_date, action_norm) → row data ─
    excel_rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        client_cell = ws.cell(row=row_idx, column=2)
        date_cell = ws.cell(row=row_idx, column=3)
        action_cell = ws.cell(row=row_idx, column=4)
        target_cell = ws.cell(row=row_idx, column=8)

        if not client_cell.value or not date_cell.value:
            continue

        entry_date_val = date_cell.value
        if isinstance(entry_date_val, datetime):
            entry_date_val = entry_date_val.date()
        elif isinstance(entry_date_val, str):
            entry_date_val = parse_date(entry_date_val)
        if entry_date_val is None:
            continue

        # Parse target date history from multiline cell
        target_raw = target_cell.value
        date_history: list[date] = []  # ordered oldest → newest

        if isinstance(target_raw, datetime):
            # Single date stored as datetime
            date_history = [target_raw.date()]
        elif isinstance(target_raw, str):
            parts = [p.strip() for p in target_raw.strip().split("\n") if p.strip()]
            parsed = [parse_date(p) for p in parts]
            valid = [d for d in parsed if d is not None]
            if valid:
                # Excel stores newest first → reverse to get oldest first
                date_history = list(reversed(valid))

        excel_rows.append({
            "row_idx": row_idx,
            "client": (client_cell.value or "").strip(),
            "entry_date": entry_date_val,
            "action": (action_cell.value or "").strip(),
            "date_history": date_history,  # oldest → newest
        })

    print(f"Parsed {len(excel_rows)} Excel data rows")
    rows_with_history = sum(1 for r in excel_rows if len(r["date_history"]) > 1)
    print(f"  {rows_with_history} rows have date revision history (>1 date)")

    # ── Connect to DB ─────────────────────────────────────────────────────────
    engine = create_async_engine(DATABASE_URL)
    Session = async_sessionmaker(engine, expire_on_commit=False)

    async with Session() as db:
        # Get admin user id for changelog attribution
        admin_result = await db.execute(select(User).limit(1))
        admin_user = admin_result.scalar_one_or_none()
        admin_id = admin_user.id if admin_user else None

        # Get all activities
        act_result = await db.execute(
            select(Activity).where(Activity.deleted_at == None)
        )
        activities = act_result.scalars().all()
        print(f"\nFound {len(activities)} activities in DB")

        # Build DB lookup: (client_norm, entry_date, action_norm[:80])
        db_lookup: dict[tuple, Activity] = {}
        for a in activities:
            key = (
                normalize_str(a.client_name),
                a.entry_date,
                normalize_str(a.action_item)[:80],
            )
            db_lookup[key] = a

        # ── Check existing history rows ───────────────────────────────────────
        existing_tdh = (await db.execute(select(func.count()).select_from(TargetDateHistory))).scalar()
        print(f"Existing target_date_history rows: {existing_tdh}")
        if existing_tdh > 0:
            print("WARNING: target_date_history already has rows — clearing before backfill")
            await db.execute(TargetDateHistory.__table__.delete())
            await db.flush()

        # ── Match Excel rows to DB activities and process ─────────────────────
        matched = 0
        unmatched = 0
        history_inserted = 0
        initial_updated = 0

        for excel_row in excel_rows:
            key = (
                normalize_str(excel_row["client"]),
                excel_row["entry_date"],
                normalize_str(excel_row["action"])[:80],
            )
            activity = db_lookup.get(key)

            if activity is None:
                # Try without action_item (some may have been truncated)
                for (cn, ed, _ai), act in db_lookup.items():
                    if cn == normalize_str(excel_row["client"]) and ed == excel_row["entry_date"]:
                        # Partial match on action
                        if normalize_str(excel_row["action"])[:40] == normalize_str(act.action_item)[:40]:
                            activity = act
                            break

            if activity is None:
                unmatched += 1
                if unmatched <= 5:
                    print(f"  UNMATCHED row {excel_row['row_idx']}: {excel_row['client']} | {excel_row['entry_date']} | {excel_row['action'][:50]}")
                continue

            matched += 1
            date_history = excel_row["date_history"]

            if not date_history:
                # No target date at all — skip
                continue

            # Update initial_target_closure_date to oldest date
            oldest_date = date_history[0]
            newest_date = date_history[-1]

            if activity.initial_target_closure_date != oldest_date:
                activity.initial_target_closure_date = oldest_date
                initial_updated += 1

            # If target_closure_date is None (import only captured datetime values,
            # not string dates like "30-Sep-25"), set it to the newest date
            if activity.target_closure_date is None and newest_date:
                activity.target_closure_date = newest_date

            # Insert history rows: one per date in the chain.
            # Use new_date as the approximate timestamp of the change
            # (best approximation we have from the Excel data).
            def date_to_ts(d: date, offset_minutes: int = 0) -> datetime:
                base = datetime(d.year, d.month, d.day, 9, 0, 0, tzinfo=timezone.utc)
                return base + timedelta(minutes=offset_minutes)

            # First entry: old_date=None (initial set), new_date=oldest
            tdh = TargetDateHistory(
                id=uuid.uuid4(),
                activity_id=activity.id,
                old_date=None,
                new_date=oldest_date,
                remarks="Initial target date",
                changed_by_id=admin_id,
                changed_at=date_to_ts(oldest_date, offset_minutes=0),
            )
            db.add(tdh)
            history_inserted += 1

            # Subsequent entries: old_date=prev, new_date=next
            for i in range(1, len(date_history)):
                prev_d = date_history[i - 1]
                new_d = date_history[i]
                tdh = TargetDateHistory(
                    id=uuid.uuid4(),
                    activity_id=activity.id,
                    old_date=prev_d,
                    new_date=new_d,
                    remarks="Target date revised",
                    changed_by_id=admin_id,
                    changed_at=date_to_ts(new_d, offset_minutes=i),
                )
                db.add(tdh)
                history_inserted += 1

        await db.flush()
        await db.commit()

        print(f"\n=== Results ===")
        print(f"  Matched:           {matched}")
        print(f"  Unmatched:         {unmatched}")
        print(f"  initial_tgt updated: {initial_updated}")
        print(f"  History rows inserted: {history_inserted}")

        # Verify
        final_count = (await db.execute(select(func.count()).select_from(TargetDateHistory))).scalar()
        print(f"  Final target_date_history count: {final_count}")

    await engine.dispose()
    print("\nDone.")


if __name__ == "__main__":
    asyncio.run(main())
